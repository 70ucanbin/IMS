import openpyxl as px
import os, shutil, traceback

from calendar import monthrange 

from flask import abort
from flask_login import current_user

from ims.service.comServ import getComUser

class __ExcelUtil(object):
    """Excelの共通処理

    各出力機能の共通部分を定義します。
    1.一時ファイルの作成処理
    2.Excel処理中のException処理
    3.一時ファイルを閉じて削除する処理

    :param template_path: 出力対象フォーマットの格納パス
    :param tmp_path: 一時ファイルの格納パス
    """
    def __init__(self, template_path, tmp_path):
        """一時ファイルの作成処理

        リクエスト別で重複しない一時パスをパラメータで取得します。
        テンプレートをそこにコピーし、ロードします。
        """
        self.template_path = template_path
        self.tmp_path = tmp_path
        try:
            os.makedirs(tmp_path)
            self.tmp_file = shutil.copy(template_path, tmp_path+'\\tmp.xlsx')
            self.book = px.load_workbook(self.tmp_file)

        except:
            #ファイルを読み込み時のエラー処理を記述
            if os.path.exists(tmp_path):
                shutil.rmtree(self.tmp_path)
            abort(500)

    def edit_file(self):
        """業務別の帳票への書き込み処理を定義します。
        
        各業務でこれをoverwriteして使用してください。
        """
        pass

    def close_file(self):
        """一時ファイルを閉じて削除する処理
        """
        try:
            self.book.close()
            shutil.rmtree(self.tmp_path)
        except:
            #ファイルを閉じる時のエラー処理を記述
            traceback.print_exc()
            abort(500)


class travelExpenses_excel(__ExcelUtil):
    """旅費精算の帳票出力処理

    :param template_path: 出力対象フォーマットの格納パス
    :param tmp_path: 一時ファイルの格納パス
    """
    def edit_file(self, userId, models):
        """帳票書き込み処理

        :param userId: 出力対象ユーザID
        :param models: 出力詳細データ格納model
        """
        try:
            if models:
                dto = getComUser(userId)
                sheet = self.book.active
                # 氏名
                sheet.cell(6, 9).value = dto.user_name
                # 期間及び提出日
                __, lastDay = monthrange(models[0].entry_year, models[0].entry_month)
                sheet.cell(8, 2).value = str(models[0].entry_year) + '年' + str(models[0].entry_month) + '月' + '1日'
                sheet.cell(8, 5).value = str(models[0].entry_year) + '年' + str(models[0].entry_month) + '月' + str(lastDay) + '日'
                sheet.cell(8, 9).value = str(models[0].entry_year) + '年' + str(models[0].entry_month) + '月' + str(lastDay) + '日'
                row = 12
                column = 1
                for model in models:
                    sheet.cell(row, column).value = model.expense_date
                    sheet.cell(row, column+1).value = model.expense_item
                    sheet.cell(row, column+2).value = model.route
                    sheet.cell(row, column+4).value = model.transit
                    sheet.cell(row, column+5).value = model.payment
                    sheet.cell(row, column+6).value = ''
                    sheet.cell(row, column+7).value = model.file_name
                    sheet.cell(row, column+8).value = model.note
                    row +=1
                self.book.save(self.tmp_file)
                result_file = open(self.tmp_file, "rb").read()
            else:
                result_file = None
        except TypeError:
            # 変なデータが登録されない限り、TypeErrorは起こらないが、念のため
            traceback.print_exc()
        super().close_file()

        return result_file


class monthly_report_excel(__ExcelUtil):
    """月報の帳票出力処理

    :param template_path: 出力対象フォーマットの格納パス
    :param tmp_path: 一時ファイルの格納パス
    """
    def edit_file(self, userId, models):
        """帳票書き込み処理

        :param userId: 出力対象ユーザID
        :param models: 出力詳細データ格納model
        """
        try:
            if models:
                dto = getComUser(userId)
                sheet = self.book.active

                # ヘッダ項目：作業者名
                sheet.cell(1, 11).value = dto.user_name
                # ヘッダ項目：年・月

                self.book.save(self.tmp_file)
                result_file = open(self.tmp_file, "rb").read()
            else:
                result_file = None
        except TypeError:
            # 変なデータが登録されない限り、TypeErrorは起こらないが、念のため
            traceback.print_exc()
        super().close_file()

        return result_file




